from flask import Blueprint, render_template, request, flash, redirect, url_for, session, jsonify
from routes.utils import (
    require_auth, get_attr_value, check_permission,
    require_permission, require_api_permission,
    get_user_by_dn, get_read_connection
)
from common import (
    load_config, get_user_by_samaccountname, get_group_by_name,
    get_service_account_connection, search_groups_for_user_addition,
    load_group_schedules, save_group_schedules
)
from forms.groups import GroupSearchForm, CreateScheduleForm, ManageMemberForm
from flask_wtf import FlaskForm
import ldap3
from ldap3 import BASE
from ldap3.utils.conv import escape_filter_chars
import logging
import re
import uuid
from datetime import datetime, date, timedelta

groups_bp = Blueprint('groups', __name__)

def sync_zimbra_member_realtime(group_name, user_sam, action):
    """
    Sincroniza um membro em tempo real com o Zimbra em segundo plano (thread) se a integração estiver ativa.
    action: 'add' ou 'remove'
    """
    import threading
    
    def _run_sync():
        try:
            from common import load_config
            config = load_config()
            if not config.get('ZIMBRA_ENABLED', False):
                return
                
            zimbra_url = config.get('ZIMBRA_API_URL')
            zimbra_user = config.get('ZIMBRA_ADMIN_USER')
            zimbra_password = config.get('ZIMBRA_ADMIN_PASSWORD')
            
            if not zimbra_url or not zimbra_user:
                return
                
            # Verifica se o grupo está mapeado
            from routes.zimbra import load_zimbra_mappings
            mappings = load_zimbra_mappings()
            zimbra_email = None
            for m in mappings:
                if m['ad_group_name'] == group_name:
                    zimbra_email = m['zimbra_dl_email']
                    break
                    
            if not zimbra_email:
                return
                
            # Busca o e-mail do usuário no AD
            from common import get_service_account_connection, get_user_by_samaccountname, get_attr_value
            conn = get_service_account_connection()
            user = get_user_by_samaccountname(conn, user_sam, ['mail', 'userPrincipalName'])
            if not user:
                return
                
            member_email = get_attr_value(user, 'mail') or get_attr_value(user, 'userPrincipalName')
            if not member_email:
                return
                
            # Conecta e atualiza o Zimbra
            from routes.zimbra_api import ZimbraSOAPClient
            client = ZimbraSOAPClient(zimbra_url, zimbra_user, zimbra_password)
            
            if action == 'add':
                client.add_dl_member(zimbra_email, member_email)
                logging.info(f"[ZIMBRA-REALTIME-SYNC] Membro '{member_email}' adicionado ao grupo Zimbra '{zimbra_email}' em tempo real.")
            elif action == 'remove':
                client.remove_dl_member(zimbra_email, member_email)
                logging.info(f"[ZIMBRA-REALTIME-SYNC] Membro '{member_email}' removido do grupo Zimbra '{zimbra_email}' em tempo real.")
                
        except Exception as e:
            logging.error(f"[ZIMBRA-REALTIME-SYNC] Erro ao sincronizar em tempo real ({action}) para o grupo {group_name}: {e}")

    # Dispara a sincronização em uma thread separada para não travar a requisição da página do usuário
    threading.Thread(target=_run_sync, daemon=True).start()

@groups_bp.route('/group_management', methods=['GET', 'POST'])
@require_auth
@require_permission(action='can_manage_groups')
def group_management():
    form = GroupSearchForm()
    groups = []
    if form.validate_on_submit():
        try:
            conn = get_service_account_connection()
            config = load_config()
            search_base = config.get('AD_SEARCH_BASE')
            query = form.search_query.data
            search_filter = f"(&(objectClass=group)(cn=*{escape_filter_chars(query)}*))"
            conn.search(search_base, search_filter, attributes=['cn', 'description', 'member', 'groupType'])
            
            groups = []
            for g in conn.entries:
                g_type_val = g.groupType.value if 'groupType' in g and g.groupType.value else 0
                
                # Active Directory groupType bitmask parsing
                is_security = (g_type_val & 2147483648) or (g_type_val < 0)
                type_str = "Segurança" if is_security else "Distribuição"
                
                scope_str = "Desconhecido"
                if g_type_val & 2:
                    scope_str = "Global"
                elif g_type_val & 4:
                    scope_str = "Domínio Local"
                elif g_type_val & 8:
                    scope_str = "Universal"
                    
                groups.append({
                    'cn': g.cn.value if 'cn' in g else 'Desconhecido',
                    'member_count': len(g.member.values) if 'member' in g and g.member.values else 0,
                    'type_scope': f"{type_str} - {scope_str}"
                })
            if not groups:
                flash(f"Nenhum grupo encontrado com o nome '{query}'.", "info")
        except Exception as e:
            flash(f"Erro ao buscar grupos: {e}", "error")
            logging.error(f"Erro ao buscar grupos com a query '{form.search_query.data}': {e}", exc_info=True)
    return render_template('manage_groups.html', form=form, groups=groups)

@groups_bp.route('/view_group/<group_name>')
@require_auth
@require_permission(action='can_manage_groups')
def view_group(group_name):
    form = FlaskForm()
    try:
        conn = get_service_account_connection()
        group = get_group_by_name(conn, group_name, attributes=['cn', 'description', 'groupType', 'mail', 'proxyAddresses'])
        if not group:
            flash(f"Grupo '{group_name}' não encontrado.", 'error')
            return redirect(url_for('groups.group_management'))
            
        g_type_val = group.groupType.value if 'groupType' in group and group.groupType.value else 0
        is_security = (g_type_val & 2147483648) or (g_type_val < 0)
        
        scope_val = 'domain_local' if (g_type_val & 4) else ('global' if (g_type_val & 2) else ('universal' if (g_type_val & 8) else 'unknown'))
        
        group_info = {
            'cn': group.cn.value if 'cn' in group else group_name,
            'description': group.description.value if 'description' in group else '',
            'is_security': is_security,
            'scope': scope_val,
            'mail': group.mail.value if 'mail' in group else '',
            'proxyAddresses': group.proxyAddresses.values if 'proxyAddresses' in group else []
        }
            
        return render_template('view_group.html', group=group_info, form=form)
    except Exception as e:
        flash(f"Erro ao carregar a página do grupo: {e}", "error")
        logging.error(f"Erro ao carregar a view do grupo '{group_name}': {e}", exc_info=True)
        return redirect(url_for('groups.group_management'))

@groups_bp.route('/add_member/<group_name>', methods=['POST'])
@require_auth
@require_permission(action='can_manage_groups')
def add_member(group_name):
    try:
        user_sam = request.form.get('user_sam')
        if not user_sam:
            flash("Login do usuário não fornecido.", 'error')
            return redirect(url_for('groups.view_group', group_name=group_name))
        conn = get_service_account_connection()
        user_to_add = get_user_by_samaccountname(conn, user_sam, ['distinguishedName'])
        group_to_modify = get_group_by_name(conn, group_name, ['distinguishedName'])
        if user_to_add and group_to_modify:
            conn.extend.microsoft.add_members_to_groups([user_to_add.distinguishedName.value], group_to_modify.distinguishedName.value)
            if conn.result['description'] == 'success':
                flash(f"Usuário '{user_sam}' adicionado ao grupo '{group_name}' com sucesso.", 'success')
                logging.info(f"[ALTERAÇÃO] Usuário '{user_sam}' adicionado permanentemente ao grupo '{group_name}' por '{session.get('ad_user')}'.")
                sync_zimbra_member_realtime(group_name, user_sam, 'add')
                try:
                    schedules = load_group_schedules()
                    schedules_to_keep = [s for s in schedules if not (s.get('user_sam') == user_sam and s.get('group_name') == group_name)]
                    if len(schedules_to_keep) < len(schedules):
                        save_group_schedules(schedules_to_keep)
                        logging.info(f"Agendamentos pendentes para '{user_sam}' no grupo '{group_name}' foram removidos devido à adição permanente.")
                except Exception as e:
                    logging.error(f"Erro ao limpar agendamentos para '{user_sam}' no grupo '{group_name}': {e}")
            else:
                flash(f"Falha ao adicionar usuário: {conn.result['message']}", 'error')
        else:
            flash("Usuário ou grupo não encontrado.", 'error')
    except Exception as e:
        flash(f"Erro ao adicionar usuário ao grupo: {e}", "error")
        logging.error(f"Erro ao adicionar usuário '{user_sam}' ao grupo '{group_name}': {e}", exc_info=True)
    return redirect(url_for('groups.view_group', group_name=group_name))

@groups_bp.route('/remove_member/<group_name>/<user_sam>', methods=['POST'])
@require_auth
@require_permission(action='can_manage_groups')
def remove_member(group_name, user_sam):
    try:
        conn = get_service_account_connection()
        user_to_remove = get_user_by_samaccountname(conn, user_sam, ['distinguishedName'])
        group_to_modify = get_group_by_name(conn, group_name, ['distinguishedName'])
        if user_to_remove and group_to_modify:
            conn.extend.microsoft.remove_members_from_groups([user_to_remove.distinguishedName.value], group_to_modify.distinguishedName.value)
            if conn.result['description'] == 'success':
                flash(f"Usuário '{user_sam}' removido do grupo '{group_name}' com sucesso.", 'success')
                logging.info(f"[ALTERAÇÃO] Usuário '{user_sam}' removido permanentemente do grupo '{group_name}' por '{session.get('ad_user')}'.")
                sync_zimbra_member_realtime(group_name, user_sam, 'remove')
                try:
                    schedules = load_group_schedules()
                    schedules_to_keep = [s for s in schedules if not (s.get('user_sam') == user_sam and s.get('group_name') == group_name)]
                    if len(schedules_to_keep) < len(schedules):
                        save_group_schedules(schedules_to_keep)
                        logging.info(f"Agendamentos pendentes para '{user_sam}' no grupo '{group_name}' foram removidos devido à remoção permanente.")
                except Exception as e:
                    logging.error(f"Erro ao limpar agendamentos para '{user_sam}' no grupo '{group_name}': {e}")
            else:
                flash(f"Falha ao remover usuário: {conn.result['message']}", 'error')
        else:
            flash("Usuário ou grupo não encontrado.", 'error')
    except Exception as e:
        flash(f"Erro ao remover usuário do grupo: {e}", "error")
        logging.error(f"Erro ao remover usuário '{user_sam}' do grupo '{group_name}': {e}", exc_info=True)
    return redirect(url_for('groups.view_group', group_name=group_name))


# API ENDPOINTS
@groups_bp.route('/api/group_members/<group_name>')
@require_auth
@require_api_permission(action='can_manage_groups')
def api_group_members(group_name):
    try:
        conn = get_service_account_connection()
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 10, type=int)
        search_query = request.args.get('query', '', type=str).strip()
        group = get_group_by_name(conn, group_name, attributes=['member'])
        if not group:
            return jsonify({'error': 'Group not found'}), 404
        member_dns = group.member.values if 'member' in group and group.member.values else []
        if search_query:
            escaped_query = re.escape(search_query)
            member_dns = [dn for dn in member_dns if re.search(f'CN={escaped_query}', dn, re.IGNORECASE)]
        total_members = len(member_dns)
        start = (page - 1) * per_page
        end = start + per_page
        paginated_dns = sorted(member_dns)[start:end]
        members_details = []
        attributes_to_get = ['displayName', 'sAMAccountName', 'title', 'l']
        for dn in paginated_dns:
            user_entry = get_user_by_dn(conn, dn, attributes=attributes_to_get)
            if user_entry:
                members_details.append({
                    'displayName': get_attr_value(user_entry, 'displayName'),
                    'sAMAccountName': get_attr_value(user_entry, 'sAMAccountName'),
                    'title': get_attr_value(user_entry, 'title'),
                    'city': get_attr_value(user_entry, 'l')
                })
            else:
                cn_part = dn.split(',')[0]
                display_name = cn_part.split('=')[1] if '=' in cn_part else cn_part
                members_details.append({
                    'displayName': f"{display_name} (Objeto desconhecido)",
                    'sAMAccountName': 'N/A', 'title': 'N/A', 'city': 'N/A'
                })
        return jsonify({
            'members': members_details,
            'total': total_members, 'page': page, 'per_page': per_page,
            'total_pages': (total_members + per_page - 1) // per_page
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@groups_bp.route('/api/batch_add_members/<group_name>', methods=['POST'])
@require_auth
@require_api_permission(action='can_manage_groups')
def api_batch_add_members(group_name):
    import os
    if 'file' not in request.files:
        return jsonify({'error': 'Nenhum arquivo enviado.'}), 400
        
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'Arquivo sem nome.'}), 400
        
    # Validar extensão do arquivo
    allowed_extensions = {'.txt', '.csv', '.xlsx'}
    _, ext = os.path.splitext(file.filename.lower())
    if ext not in allowed_extensions:
        return jsonify({'error': 'Formato de arquivo não suportado. Use apenas .txt, .csv ou .xlsx.'}), 400
        
    sam_accounts = []
    try:
        if ext == '.xlsx':
            from openpyxl import load_workbook
            wb = load_workbook(file, read_only=True)
            sheet = wb.active
            for row in sheet.iter_rows(values_only=True):
                if row and row[0]:
                    val = str(row[0]).strip()
                    if val and val.lower() not in ['login', 'samaccountname', 'usuario', 'usuário', 'username', 'email', 'e-mail', 'mail']:
                        sam_accounts.append(val)
        else: # .txt ou .csv
            content = file.read().decode('utf-8', errors='ignore')
            # Divide o conteúdo por novas linhas, vírgulas ou ponto-e-vírgula
            for line in re.split(r'[\r\n,;]+', content):
                val = line.strip()
                if val:
                    if val.lower() not in ['login', 'samaccountname', 'usuario', 'usuário', 'username', 'email', 'e-mail', 'mail']:
                        sam_accounts.append(val)
    except Exception as e:
        logging.error(f"Erro ao ler arquivo de lote para o grupo {group_name}: {e}", exc_info=True)
        return jsonify({'error': f"Erro ao ler arquivo: {str(e)}"}), 500

    # Remove duplicados da lista mantendo a ordem original
    sam_accounts = list(dict.fromkeys(sam_accounts))
    
    if not sam_accounts:
        return jsonify({'error': 'Nenhum login ou e-mail válido encontrado no arquivo.'}), 400
        
    results = []
    stats = {
        'total': len(sam_accounts),
        'success': 0,
        'already_member': 0,
        'failed': 0
    }
    
    def get_user_by_login_or_email(conn, identifier, attributes=None):
        if attributes is None:
            attributes = ['distinguishedName', 'displayName', 'sAMAccountName']
        config = load_config()
        search_base = config.get('AD_SEARCH_BASE')
        if not search_base:
            if conn.server.info and conn.server.info.other:
                search_base = conn.server.info.other['defaultNamingContext'][0]
            else:
                raise Exception("AD_SEARCH_BASE não configurado.")
        
        # Se contiver '@', busca por mail ou userPrincipalName. Caso contrário, busca por sAMAccountName
        if '@' in identifier:
            search_filter = f"(|(mail={escape_filter_chars(identifier)})(userPrincipalName={escape_filter_chars(identifier)}))"
        else:
            search_filter = f"(sAMAccountName={escape_filter_chars(identifier)})"
            
        conn.search(search_base, search_filter, attributes=attributes)
        if conn.entries:
            return conn.entries[0]
        return None

    try:
        conn = get_service_account_connection()
        group_to_modify = get_group_by_name(conn, group_name, ['distinguishedName', 'member'])
        if not group_to_modify:
            return jsonify({'error': 'Grupo não encontrado.'}), 404
            
        group_dn = group_to_modify.distinguishedName.value
        current_member_dns = set(group_to_modify.member.values) if 'member' in group_to_modify and group_to_modify.member.values else set()
        
        for sam in sam_accounts:
            try:
                # Busca o usuário por login ou e-mail/UPN
                user = get_user_by_login_or_email(conn, sam, ['distinguishedName', 'displayName', 'sAMAccountName'])
                if not user:
                    results.append({
                        'sam': sam,
                        'name': 'Desconhecido',
                        'status': 'error',
                        'message': 'Não encontrado no AD'
                    })
                    stats['failed'] += 1
                    continue
                    
                user_dn = user.distinguishedName.value
                display_name = get_attr_value(user, 'displayName') or sam
                sam_real = user.sAMAccountName.value if 'sAMAccountName' in user else sam
                
                # Verifica se já é membro
                if user_dn in current_member_dns:
                    results.append({
                        'sam': sam,
                        'name': display_name,
                        'status': 'already_member',
                        'message': f"Já é membro (Login: {sam_real})"
                    })
                    stats['already_member'] += 1
                    continue
                    
                # Adiciona ao grupo
                conn.extend.microsoft.add_members_to_groups([user_dn], group_dn)
                if conn.result['description'] == 'success':
                    results.append({
                        'sam': sam,
                        'name': display_name,
                        'status': 'success',
                        'message': f"Adicionado (Login: {sam_real})"
                    })
                    stats['success'] += 1
                    logging.info(f"[LOTE-ALTERAÇÃO] Usuário '{sam_real}' (pesquisado por '{sam}') adicionado ao grupo '{group_name}' em lote por '{session.get('user_display_name')}'.")
                    sync_zimbra_member_realtime(group_name, sam_real, 'add')
                    
                    # Tenta remover de agendamentos pendentes se houver
                    try:
                        schedules = load_group_schedules()
                        schedules_to_keep = [s for s in schedules if not (s.get('user_sam') == sam_real and s.get('group_name') == group_name)]
                        if len(schedules_to_keep) < len(schedules):
                            save_group_schedules(schedules_to_keep)
                    except Exception as e_sched:
                        logging.error(f"Erro ao limpar agendamento em lote para {sam_real} no grupo {group_name}: {e_sched}")
                else:
                    results.append({
                        'sam': sam,
                        'name': display_name,
                        'status': 'error',
                        'message': f"Erro AD: {conn.result['message']}"
                    })
                    stats['failed'] += 1
            except Exception as e_user:
                results.append({
                    'sam': sam,
                    'name': 'Desconhecido',
                    'status': 'error',
                    'message': f"Erro: {str(e_user)}"
                })
                stats['failed'] += 1
                logging.error(f"Erro ao adicionar usuário {sam} em lote ao grupo {group_name}: {e_user}", exc_info=True)
                
        return jsonify({
            'success': True,
            'stats': stats,
            'results': results
        })
    except Exception as e:
        logging.error(f"Erro crítico no processamento em lote para o grupo {group_name}: {e}", exc_info=True)
        return jsonify({'error': f"Erro interno do servidor: {str(e)}"}), 500

@groups_bp.route('/api/user_groups/<username>')
@require_auth
def api_user_groups(username):
    try:
        conn = get_read_connection()
        user = get_user_by_samaccountname(conn, username, attributes=['memberOf'])
        if not user:
            return jsonify([])
        group_dns = user.memberOf.values if 'memberOf' in user and user.memberOf.values else []
        groups_details = []
        for dn in group_dns:
            conn.search(dn, '(objectClass=group)', BASE, attributes=['cn', 'description'])
            if conn.entries:
                group_entry = conn.entries[0]
                groups_details.append({
                    'cn': get_attr_value(group_entry, 'cn'),
                    'description': get_attr_value(group_entry, 'description')
                })
        sorted_groups = sorted(groups_details, key=lambda g: g.get('cn', '').lower())
        return jsonify(sorted_groups)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@groups_bp.route('/api/search_groups', methods=['GET'])
@require_auth
@require_api_permission(action='can_manage_groups')
def api_search_groups():
    query = request.args.get('q', '')
    username = request.args.get('username', '')
    if not query or len(query) < 3:
        return jsonify([])
    try:
        conn = get_service_account_connection()
        groups = search_groups_for_user_addition(conn, query, username)
        return jsonify(groups)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@groups_bp.route('/api/add_user_to_group', methods=['POST'])
@require_auth
@require_api_permission(action='can_manage_groups')
def api_add_user_to_group():
    data = request.get_json()
    username = data.get('username')
    group_name = data.get('group_name')
    if not username or not group_name:
        return jsonify({'error': 'Nome de usuário e nome do grupo são obrigatórios.'}), 400
    try:
        conn = get_service_account_connection()
        user = get_user_by_samaccountname(conn, username, ['distinguishedName'])
        group = get_group_by_name(conn, group_name, ['distinguishedName'])
        if not user or not group:
            return jsonify({'error': 'Usuário ou grupo não encontrado.'}), 404
        conn.extend.microsoft.add_members_to_groups([user.distinguishedName.value], group.distinguishedName.value)
        if conn.result['description'] == 'success':
            logging.info(f"[ALTERAÇÃO] Usuário '{username}' adicionado permanentemente ao grupo '{group_name}' por '{session.get('user_display_name')}'.")
            sync_zimbra_member_realtime(group_name, username, 'add')
            try:
                schedules = load_group_schedules()
                schedules_to_keep = [s for s in schedules if not (s.get('user_sam') == username and s.get('group_name') == group_name)]
                if len(schedules_to_keep) < len(schedules):
                    save_group_schedules(schedules_to_keep)
            except Exception as e:
                logging.error(f"Erro ao limpar agendamentos para '{username}' no grupo '{group_name}': {e}")
            return jsonify({'success': True, 'message': 'Usuário adicionado ao grupo com sucesso.'})
        else:
            raise Exception(f"Falha do LDAP: {conn.result['message']}")
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@groups_bp.route('/api/remove_user_from_group', methods=['POST'])
@require_auth
@require_api_permission(action='can_manage_groups')
def api_remove_user_from_group():
    data = request.get_json()
    username = data.get('username')
    group_name = data.get('group_name')
    if not username or not group_name:
        return jsonify({'error': 'Nome de usuário e nome do grupo são obrigatórios.'}), 400
    try:
        conn = get_service_account_connection()
        user = get_user_by_samaccountname(conn, username, ['distinguishedName'])
        group = get_group_by_name(conn, group_name, ['distinguishedName'])
        if not user or not group:
            return jsonify({'error': 'Usuário ou grupo não encontrado.'}), 404
        conn.extend.microsoft.remove_members_from_groups([user.distinguishedName.value], group.distinguishedName.value)
        if conn.result['description'] == 'success':
            logging.info(f"[ALTERAÇÃO] Usuário '{username}' removido permanentemente do grupo '{group_name}' por '{session.get('user_display_name')}'.")
            sync_zimbra_member_realtime(group_name, username, 'remove')
            try:
                schedules = load_group_schedules()
                schedules_to_keep = [s for s in schedules if not (s.get('user_sam') == username and s.get('group_name') == group_name)]
                if len(schedules_to_keep) < len(schedules):
                    save_group_schedules(schedules_to_keep)
            except Exception as e:
                logging.error(f"Erro ao limpar agendamentos para '{username}' no grupo '{group_name}': {e}")
            return jsonify({'success': True, 'message': 'Usuário removido do grupo com sucesso.'})
        else:
            raise Exception(f"Falha do LDAP: {conn.result['message']}")
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@groups_bp.route('/api/remove_users_from_group', methods=['POST'])
@require_auth
@require_api_permission(action='can_manage_groups')
def api_remove_users_from_group():
    data = request.get_json()
    usernames = data.get('usernames', [])
    group_name = data.get('group_name')
    
    if not usernames or not group_name:
        return jsonify({'error': 'Nomes de usuários e nome do grupo são obrigatórios.'}), 400
        
    try:
        conn = get_service_account_connection()
        group = get_group_by_name(conn, group_name, ['distinguishedName'])
        if not group:
            return jsonify({'error': 'Grupo não encontrado.'}), 404
            
        group_dn = group.distinguishedName.value
        
        results = []
        stats = {'total': len(usernames), 'success': 0, 'failed': 0}
        
        for username in usernames:
            try:
                user = get_user_by_samaccountname(conn, username, ['distinguishedName'])
                if not user:
                    results.append({'sam': username, 'status': 'error', 'message': 'Não encontrado no AD'})
                    stats['failed'] += 1
                    continue
                    
                user_dn = user.distinguishedName.value
                conn.extend.microsoft.remove_members_from_groups([user_dn], group_dn)
                if conn.result['description'] == 'success':
                    results.append({'sam': username, 'status': 'success', 'message': 'Removido com sucesso'})
                    stats['success'] += 1
                    logging.info(f"[LOTE-REMOÇÃO] Usuário '{username}' removido do grupo '{group_name}' em lote por '{session.get('user_display_name')}'.")
                    sync_zimbra_member_realtime(group_name, username, 'remove')
                    
                    # Limpar agendamentos se houver
                    try:
                        schedules = load_group_schedules()
                        schedules_to_keep = [s for s in schedules if not (s.get('user_sam') == username and s.get('group_name') == group_name)]
                        if len(schedules_to_keep) < len(schedules):
                            save_group_schedules(schedules_to_keep)
                    except Exception as e_sched:
                        logging.error(f"Erro ao limpar agendamento na remoção em lote para {username}: {e_sched}")
                else:
                    results.append({'sam': username, 'status': 'error', 'message': f"Erro AD: {conn.result['message']}"})
                    stats['failed'] += 1
            except Exception as e_user:
                results.append({'sam': username, 'status': 'error', 'message': str(e_user)})
                stats['failed'] += 1
                
        return jsonify({
            'success': True,
            'stats': stats,
            'results': results
        })
    except Exception as e:
        logging.error(f"Erro crítico na remoção múltipla do grupo {group_name}: {e}", exc_info=True)
        return jsonify({'error': f"Erro interno do servidor: {str(e)}"}), 500

@groups_bp.route('/api/add_user_to_group_temp', methods=['POST'])
@require_auth
@require_api_permission(action='can_manage_groups')
def api_add_user_to_group_temp():
    data = request.get_json()
    username = data.get('username')
    group_name = data.get('group_name')
    start_date_str = data.get('start_date')
    end_date_str = data.get('end_date')
    if not all([username, group_name, start_date_str, end_date_str]):
        return jsonify({'error': 'Todos os campos são obrigatórios.'}), 400
    try:
        start_date = date.fromisoformat(start_date_str)
        end_date = date.fromisoformat(end_date_str)
        today = date.today()
        if start_date > end_date:
            return jsonify({'error': 'A data de início não pode ser posterior à data de fim.'}), 400
        conn = get_service_account_connection()
        user_to_add = get_user_by_samaccountname(conn, username, ['distinguishedName'])
        group_to_modify = get_group_by_name(conn, group_name, ['distinguishedName'])
        if not user_to_add or not group_to_modify:
            return jsonify({'error': 'Usuário ou grupo não encontrado.'}), 404
        schedules = load_group_schedules()
        schedule_id = str(uuid.uuid4())
        add_schedule = {'id': schedule_id, 'user_sam': username.lower(), 'group_name': group_name, 'action': 'add', 'execution_date': start_date.isoformat()}
        schedules.append(add_schedule)
        if start_date <= today:
            conn.extend.microsoft.add_members_to_groups([user_to_add.distinguishedName.value], group_to_modify.distinguishedName.value)
            if conn.result['description'] != 'success':
                schedules.pop()
                raise Exception(f"Falha do LDAP: {conn.result['message']}")
        remove_schedule = {'id': schedule_id, 'user_sam': username.lower(), 'group_name': group_name, 'action': 'remove', 'execution_date': end_date.isoformat()}
        schedules.append(remove_schedule)
        save_group_schedules(schedules)
        return jsonify({'success': True, 'message': 'Agendamento realizado com sucesso.'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@groups_bp.route('/api/remove_group_schedule', methods=['POST'])
@require_auth
@require_api_permission(action='can_manage_groups')
def api_remove_group_schedule():
    data = request.get_json()
    user_sam = data.get('user_sam')
    group_name = data.get('group_name')
    remove_now = data.get('remove_now', False)
    try:
        schedules = load_group_schedules()
        new_schedules = [s for s in schedules if not (s['user_sam'].lower() == user_sam.lower() and s['group_name'] == group_name)]
        if len(new_schedules) == len(schedules):
            return jsonify({'error': 'Agendamento não encontrado.'}), 404
        save_group_schedules(new_schedules)
        msg = "Agendamento removido."
        if remove_now:
            conn = get_service_account_connection()
            user = get_user_by_samaccountname(conn, user_sam, ['distinguishedName'])
            group = get_group_by_name(conn, group_name, ['distinguishedName'])
            if user and group:
                conn.extend.microsoft.remove_members_from_groups([user.distinguishedName.value], group.distinguishedName.value)
                msg += " Usuário removido do grupo imediatamente."
        return jsonify({'success': True, 'message': msg})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@groups_bp.route('/api/search_users_for_group/<group_name>')
@require_auth
@require_api_permission(action='can_manage_groups')
def api_search_users_for_group(group_name):
    query = request.args.get('query', '')
    if len(query) < 3:
        return jsonify([])
    try:
        conn = get_read_connection()
        from common import search_groups_for_user_addition
        # Aqui, na verdade, queremos buscar usuários que NÃO estão no grupo
        # O helper search_groups_for_user_addition faz o oposto (busca grupos para um usuário)
        # Vamos implementar a busca de usuários aqui
        config = load_config()
        search_base = config.get('AD_SEARCH_BASE')
        
        # Primeiro pegamos os membros do grupo para excluir
        group = get_group_by_name(conn, group_name, ['member'])
        current_members = set(group.member.values) if group and 'member' in group and group.member.values else set()
        
        search_filter = f"(&(objectClass=user)(objectCategory=person)(|(displayName=*{query}*)(sAMAccountName=*{query}*)))"
        conn.search(search_base, search_filter, attributes=['displayName', 'sAMAccountName', 'distinguishedName', 'title', 'l'])
        
        results = []
        for entry in conn.entries:
            if entry.distinguishedName.value not in current_members:
                results.append({
                    'displayName': entry.displayName.value or entry.sAMAccountName.value,
                    'sAMAccountName': entry.sAMAccountName.value,
                    'title': entry.title.value if 'title' in entry else 'N/A',
                    'city': entry.l.value if 'l' in entry else 'N/A'
                })
        return jsonify(results)
    except Exception as e:
        logging.error(f"Erro na busca de usuários para grupo: {e}")
        return jsonify({'error': str(e)}), 500

@groups_bp.route('/api/export_group_members/<group_name>')
@require_auth
@require_permission(action='can_export_data')
def api_export_group_members(group_name):
    import csv
    import io
    from flask import Response
    
    try:
        conn = get_read_connection()
        group = get_group_by_name(conn, group_name, ['member'])
        if not group or 'member' not in group or not group.member.values:
            return "Grupo não encontrado ou sem membros.", 404
            
        member_dns = group.member.values
        # Busca detalhes dos membros
        output = io.StringIO()
        writer = csv.writer(output, delimiter=';')
        writer.writerow(['Nome', 'Login', 'Email', 'Cargo', 'Departamento', 'Unidade'])
        
        # Busca em lotes para evitar filtros gigantes
        batch_size = 50
        for i in range(0, len(member_dns), batch_size):
            batch = member_dns[i:i+batch_size]
            ldap_filter = "(| " + "".join([f"(distinguishedName={dn})" for dn in batch]) + ")"
            conn.search(conn.server.info.other['defaultNamingContext'][0], ldap_filter, attributes=['displayName', 'sAMAccountName', 'mail', 'title', 'department', 'l'])
            for entry in conn.entries:
                writer.writerow([
                    entry.displayName.value or '',
                    entry.sAMAccountName.value or '',
                    entry.mail.value or '',
                    entry.title.value if 'title' in entry else '',
                    entry.department.value if 'department' in entry else '',
                    entry.l.value if 'l' in entry else ''
                ])
                
        output.seek(0)
        return Response(
            output.getvalue(),
            mimetype="text/csv",
            headers={"Content-disposition": f"attachment; filename=membros_{group_name}.csv"}
        )
    except Exception as e:
        logging.error(f"Erro ao exportar membros do grupo {group_name}: {e}")
        return f"Erro ao exportar: {str(e)}", 500

@groups_bp.route('/api/update_group_settings', methods=['POST'])
@require_auth
@require_api_permission(action='can_manage_groups')
def api_update_group_settings():
    data = request.get_json()
    group_name = data.get('group_name')
    is_security = data.get('is_security', True)
    scope = data.get('scope', 'global')
    
    if not group_name:
        return jsonify({'error': 'Nome do grupo obrigatório'}), 400
        
    try:
        conn = get_service_account_connection()
        group = get_group_by_name(conn, group_name, ['distinguishedName', 'groupType'])
        if not group:
            return jsonify({'error': 'Grupo não encontrado'}), 404
            
        current_type = group.groupType.value if 'groupType' in group and group.groupType.value else 0
        
        # Limpar apenas bit de segurança (0x80000000)
        new_type = current_type & ~2147483648
        
        # Adicionar bit de segurança se for o caso
        if is_security:
            new_type = new_type | 2147483648
            
        conn.modify(group.distinguishedName.value, {'groupType': [(ldap3.MODIFY_REPLACE, [new_type])]})
        
        if conn.result['description'] == 'success':
            logging.info(f"Tipo do grupo '{group_name}' alterado por '{session.get('user_display_name')}'.")
            return jsonify({'success': True, 'message': 'Configurações atualizadas'})
        else:
            return jsonify({'error': f"Falha no LDAP: {conn.result['message']}"}), 400
            
    except Exception as e:
        logging.error(f"Erro ao atualizar configurações do grupo {group_name}: {e}")
        return jsonify({'error': str(e)}), 500

@groups_bp.route('/api/group_exchange_info/<group_name>')
@require_auth
@require_api_permission(action='can_manage_groups')
def api_group_exchange_info(group_name):
    try:
        conn = get_read_connection()
        group = get_group_by_name(conn, group_name, attributes=['proxyAddresses', 'mail'])
        if not group:
            return jsonify({'error': 'Grupo não encontrado'}), 404
            
        primary_email = group.mail.value if 'mail' in group else None
        proxy_addresses = group.proxyAddresses.values if 'proxyAddresses' in group else []
        
        if not primary_email:
            for p in proxy_addresses:
                if str(p).startswith('SMTP:'):
                    primary_email = str(p)[5:]
                    break
                    
        return jsonify({
            'primary_email': primary_email,
            'proxy_addresses': [str(p) for p in proxy_addresses if str(p).lower().startswith('smtp:')]
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@groups_bp.route('/api/add_group_proxy_address', methods=['POST'])
@require_auth
@require_api_permission(action='can_manage_groups')
def api_add_group_proxy_address():
    data = request.get_json()
    group_name = data.get('group_name')
    new_proxy = data.get('proxy_address')
    
    if not group_name or not new_proxy:
        return jsonify({'error': 'Grupo e alias são obrigatórios'}), 400
        
    try:
        if not new_proxy.lower().startswith('smtp:'):
            new_proxy = f"smtp:{new_proxy}"
            
        conn = get_service_account_connection()
        group = get_group_by_name(conn, group_name, attributes=['proxyAddresses', 'distinguishedName', 'mail'])
        if not group:
            return jsonify({'error': 'Grupo não encontrado'}), 404
            
        proxy_addresses = [str(p) for p in group.proxyAddresses.values] if 'proxyAddresses' in group else []
        
        if any(p.lower() == new_proxy.lower() for p in proxy_addresses):
            return jsonify({'error': 'Alias já existe'}), 400
            
        proxy_addresses.append(new_proxy)
        conn.modify(group.distinguishedName.value, {'proxyAddresses': [(ldap3.MODIFY_REPLACE, proxy_addresses)]})
        
        if conn.result['description'] == 'success':
            logging.info(f"Alias '{new_proxy}' adicionado ao grupo '{group_name}' por '{session.get('user_display_name')}'")
            return jsonify({'success': True})
        else:
            return jsonify({'error': f"LDAP error: {conn.result['message']}"}), 400
            
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@groups_bp.route('/api/remove_group_proxy_address', methods=['POST'])
@require_auth
@require_api_permission(action='can_manage_groups')
def api_remove_group_proxy_address():
    data = request.get_json()
    group_name = data.get('group_name')
    target_proxy = data.get('proxy_address')
    
    if not group_name or not target_proxy:
        return jsonify({'error': 'Grupo e alias são obrigatórios'}), 400
        
    try:
        conn = get_service_account_connection()
        group = get_group_by_name(conn, group_name, attributes=['proxyAddresses', 'distinguishedName'])
        if not group:
            return jsonify({'error': 'Grupo não encontrado'}), 404
            
        proxy_addresses = [str(p) for p in group.proxyAddresses.values] if 'proxyAddresses' in group else []
        
        if target_proxy.lower().startswith('smtp:'):
            alias_to_remove = target_proxy[5:]
        else:
            alias_to_remove = target_proxy
            
        if f"SMTP:{alias_to_remove.lower()}" in [p.upper() if p.startswith('SMTP:') else p.lower() for p in proxy_addresses]:
            return jsonify({'error': 'Não é possível remover o endereço principal. Defina outro como principal primeiro.'}), 400
            
        new_proxies = [p for p in proxy_addresses if p.lower() != f"smtp:{alias_to_remove.lower()}" and p.lower() != f"smtp:{target_proxy.lower()}"]
        
        if len(new_proxies) == len(proxy_addresses):
            return jsonify({'error': 'Alias não encontrado no grupo'}), 400
            
        conn.modify(group.distinguishedName.value, {'proxyAddresses': [(ldap3.MODIFY_REPLACE, new_proxies)]})
        
        if conn.result['description'] == 'success':
            logging.info(f"Alias removido do grupo '{group_name}' por '{session.get('user_display_name')}'")
            return jsonify({'success': True})
        else:
            return jsonify({'error': f"LDAP error: {conn.result['message']}"}), 400
            
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@groups_bp.route('/api/set_group_primary_email', methods=['POST'])
@require_auth
@require_api_permission(action='can_manage_groups')
def api_set_group_primary_email():
    data = request.get_json()
    group_name = data.get('group_name')
    new_primary = data.get('new_primary')
    
    if not group_name or not new_primary:
        return jsonify({'error': 'Grupo e novo e-mail primário são obrigatórios'}), 400
        
    try:
        if new_primary.lower().startswith('smtp:'):
            new_primary = new_primary[5:]
            
        conn = get_service_account_connection()
        group = get_group_by_name(conn, group_name, attributes=['proxyAddresses', 'distinguishedName', 'mail'])
        if not group:
            return jsonify({'error': 'Grupo não encontrado'}), 404
            
        proxy_addresses = [str(p) for p in group.proxyAddresses.values] if 'proxyAddresses' in group else []
        
        new_proxies = []
        for p in proxy_addresses:
            if p.lower().startswith('smtp:'):
                clean_p = p[5:]
                if clean_p.lower() == new_primary.lower():
                    new_proxies.append(f"SMTP:{clean_p}")
                else:
                    new_proxies.append(f"smtp:{clean_p}")
            else:
                new_proxies.append(p)
                
        if not any(p.startswith('SMTP:') for p in new_proxies):
            new_proxies.append(f"SMTP:{new_primary}")
            
        modifications = {
            'proxyAddresses': [(ldap3.MODIFY_REPLACE, new_proxies)],
            'mail': [(ldap3.MODIFY_REPLACE, [new_primary])]
        }
        
        conn.modify(group.distinguishedName.value, modifications)
        
        if conn.result['description'] == 'success':
            logging.info(f"E-mail primário do grupo '{group_name}' alterado para '{new_primary}' por '{session.get('user_display_name')}'")
            return jsonify({'success': True})
        else:
            return jsonify({'error': f"LDAP error: {conn.result['message']}"}), 400
            
    except Exception as e:
        return jsonify({'error': str(e)}), 500
